import threading
from tkinter import ttk, messagebox
from bs4 import BeautifulSoup
from seleniumbase import Driver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import tkinter as tk
import tkinter.font as tkFont
import pandas as pd
import warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

prog_text = ''


# For inserting the LeaderBoard Names in the List
def printName(name, o):   # 'span-flex-4' | acm: 'span-flex-2 acm-leaderboard-cell'
    global prog_text
    i = 0
    atr = ''
    if o == 1:
        atr = 'span-flex-4'
    else:
        atr = 'span-flex-2 acm-leaderboard-cell'
    for a in soup.findAll("div", attrs={'class': atr}):
        k = str(a.text).strip()
        if k == 'Name' or k == 'User':
            pass
        else:
            name.append(k)
        i = i + 1
    print(name)


# For inserting the LeaderBoard Scores in the List
def printScore(score, o):  # 'span-flex-3'   | acm:  'span-flex-1 acm-leaderboard-cell'
    global prog_text
    i = 1
    atr = ''
    if o == 1:
        atr = 'span-flex-3'
    else:
        atr = 'span-flex-1 acm-leaderboard-cell'
    for a in soup.findAll("div", attrs={'class': atr}):
        if i%2 != 0:
            k = str(a.text).strip()
            if k == 'Score':
                pass
            elif o == 1:
                score.append(k)
            else:
                if k == '-':
                    score.append('0')
                else:
                    score.append(int(k)*10)
        i = i + 1
    print(score)


def generateExcelSheet(name, df):
    # Sort the DataFrame by 'Score' column in descending order
    df = df.sort_values(by='Score', ascending=False)

    # Create an Excel writer using openpyxl

    writer = pd.ExcelWriter(f'Leaderboards/{name}.xlsx', engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    worksheet = writer.sheets['Sheet1']

    # Define the cell styles
    font_header = Font(name='Arial', size=18, bold=True)
    font_body = Font(name='Arial', size=14, bold=True)
    fill_header = PatternFill(start_color='00ADEAEA', end_color='00ADEAEA', fill_type='solid')
    fill_body = PatternFill(start_color='00C7ECEC', end_color='00C7ECEC', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    border = Border(bottom=Side(style='medium'))

    df.insert(0, 'Rank', range(1, len(df) + 1))

    # Set the width of the 'Rank' column
    worksheet.column_dimensions['A'].width = 22
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 30
    # Apply the styling to the 'Rank' column
    for cell in worksheet['A']:
        cell.fill = fill_body
        cell.font = font_body
        cell.alignment = align_center
        cell.border = border

    # Apply formatting to the header row
    for col_num, value in enumerate(df.columns.values):
        worksheet.cell(row=1, column=col_num + 1).value = value
        worksheet.cell(row=1, column=col_num + 1).font = font_header
        worksheet.cell(row=1, column=col_num + 1).fill = fill_header
        worksheet.cell(row=1, column=col_num + 1).alignment = align_center
        worksheet.cell(row=1, column=col_num + 1).border = border

    # Apply formatting to the body cells
    for row_num, row in enumerate(df.values):
        for col_num, value in enumerate(row):
            worksheet.cell(row=row_num + 2, column=col_num + 1).value = value
            worksheet.cell(row=row_num + 2, column=col_num + 1).font = font_body
            worksheet.cell(row=row_num + 2, column=col_num + 1).fill = fill_body
            worksheet.cell(row=row_num + 2, column=col_num + 1).alignment = align_center
            worksheet.cell(row=row_num + 2, column=col_num + 1).border = border

    # Save the Excel file
    writer.close()


def getAll(tracker_names):
    global prog_text
    root.attributes('-disabled', True)
    progress_window = tk.Toplevel(root)
    progress_window.iconbitmap('venv/logo.ico')
    progress_window.title("Please Wait...")
    progress_window["borderwidth"] = "5px"
    progress_window["relief"] = "groove"
    progress_window.geometry("800x400")
    progress_window.resizable(False, False)
    progress_window['background'] = '#404445'
    progress_text = tk.Text(progress_window, height=30, width=80)
    progress_text['background'] = "grey"
    progress_text['fg'] = 'white'
    ft = tkFont.Font(family='Times', size=20, weight='bold')
    progress_text['font'] = ft
    print(f'progress text: {prog_text}'+ '\n')
    progress_text.config(state=tk.NORMAL)
    progress_text.insert(tk.END, prog_text+ '\n')
    progress_text.see(tk.END)
    progress_text.config(state=tk.DISABLED)
    progress_text.pack(pady=80)

    style = ttk.Style()
    style.theme_use('clam')  # Use the 'clam' theme as a base

    # Configure the style of the progress bar
    style.configure("TProgressbar",
                    thickness=20,  # Customize the thickness of the progress bar
                    troughcolor='lightgrey',  # Set the background color
                    background='#FF6C40',  # Set the color of the progress bar
                    )
    progress = ttk.Progressbar(progress_window, mode='determinate', style="TProgressbar")
    width = 800
    height = 400
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
    progress_window.geometry(alignstr)
    progress.pack(padx=10, ipady=20)
    progress.place(x=50, y=10, width=700, height=50)

    def generate_sheets_thread():
        global prog_text
        total_sheets = len(tracker_names)
        finished_sheets = 0
        progress_percent = 0
        warnings.filterwarnings('ignore')
        nameA = list()
        scoreA = list()
        dfA = pd.DataFrame({'Name': nameA, 'Score': scoreA})

        def check_close():
            nonlocal progress_window
            if progress_window.winfo_viewable():
                progress_window.after(1000, check_close)
            else:
                messagebox.showinfo("Process Interrupted", "Sheets generation process was interrupted.")
                progress_window.destroy()

        for pp in tracker_names:
            # backbone acm-challenge-balloon
            driver = Driver(headless=True, uc=True)
            s = BeautifulSoup(driver.page_source, features="html.parser")
            driver.close()
            if not s.find(class_="backbone acm-challenge-balloon"):
                pp = {1: pp}
            else:
                pp = {2: pp}
            for u, v in pp.items():
                name1 = []
                score1 = []
                df1 = pd.DataFrame({'Name': name1, 'Score': score1})
                pp = 1
                cho = 0
                while pp < 500:
                    ppp = pp
                    # driver2.options.add_experimental_option("excludeSwitches", ["enable-logging"])
                    name2 = []
                    score2 = []
                    try:
                        driver2 = Driver(headless=True, uc=True)
                        driver2.get(f'https://www.hackerrank.com/contests/{v}/leaderboard/{pp}')
                    except :
                        messagebox.showerror('Error', 'No Internet!')
                        return
                    try:
                        global element
                        element = WebDriverWait(driver2, 8).until(
                            lambda x: x.find_element(By.CLASS_NAME, "span-flex-1"))
                    except Exception:
                        pass
                    content2 = driver2.page_source
                    global soup
                    soup = BeautifulSoup(content2, features="html.parser")
                    if not soup.find(class_="leaderboard-list-view"):
                        break
                    printName(name2, u)
                    printScore(score2, u)
                    df2 = pd.DataFrame({'Name': name2, 'Score': [float(i) for i in score2]})
                    df1 = pd.concat([df2, df1], ignore_index=True)

                    if len(name2) == 0 and cho <= 10:
                        pp = ppp
                        cho = cho + 1
                    elif cho > 10:
                        break
                    driver2.close()
                    pp = pp + 1
                names1 = df1.to_dict().get("Name")
                scores1 = df1.to_dict().get("Score")
                d1 = dict()
                for i in names1:
                    d1.update({names1.get(i): float(scores1.get(i))})

                names2 = dfA.to_dict().get("Name")
                scores2 = dfA.to_dict().get("Score")
                d2 = dict()
                for i in names2:
                    d2.update({names2.get(i): float(scores2.get(i))})

                d3 = dict()
                for i in d1:
                    if i in d2:
                        d3.update({i: float(d1.get(i)) + float(d2.get(i))})
                    else:
                        d3.update({i: float(d1.get(i))})

                for i in d2:
                    if i in d1:
                        continue
                    else:
                        d3.update({i: float(d2.get(i))})

                namesf = []
                scoref = []
                for i, uu in d3.items():
                    namesf.append(i)
                    scoref.append(float(uu))

                finished_sheets += 1
                print(f'Finished : {v}\n')
                for l, k in enumerate(namesf):
                    prog_text += str(k)+":"+str(int(l))+'\n'
                prog_text += f'\nFinished {v}!\n'
                print(f' Inside prog_text :  {prog_text}')
                progress_text.config(state=tk.NORMAL)
                progress_text.insert(tk.END, prog_text+ '\n')
                progress_text.see(tk.END)
                progress_text.config(state=tk.DISABLED)
                progress_percent = int(finished_sheets / total_sheets * 100)
                progress['value'] = progress_percent
                progress_window.update()
                dfA = pd.DataFrame({"Name": namesf, "Score": scoref})
                generateExcelSheet(v, df1)
        generateExcelSheet("TotalHackerrankLeaderboard", dfA)

        if progress_window.winfo_viewable():
            messagebox.showinfo("Process Completed", "Sheets generated successfully.")
            root.attributes('-disabled', False)
            progress_window.wm_protocol(name='WM_DELETE_WINDOW')
            progress_window.destroy()
            root.state = 'normal'
    threading.Thread(target=generate_sheets_thread).start()
    prog_text = ''


def on_closing():
    root.destroy()

def generate_sheets(ids):
    getAll(ids)


def GButton_486_command():
    global prog_text
    inp = entry.get(1.0, 'end-1c')
    if inp == '   Enter Comma Separated values of HACKERRANK_CONTEST_ID\'s':
        messagebox.showerror('Error', 'Enter Something!')
        return
    try :
        inp = inp.replace(' ','').replace('\n','').split(',')
        d = list()
        prog_text += 'Generating leaderboards for : ('
        for i in inp:
            d.append(i)
            prog_text += str(i + ",")
        prog_text = prog_text[:-1]+ ')\n'
        print(d)
        generate_sheets(d)
    except :
        messagebox.showerror('Error', 'Something Went Wrong!')
        return


# Create the main window
root = tk.Tk()
root.title("Hackerrank Leaderboard")
root.configure(background='#404445')
width = 1142
height = 697
screenwidth = root.winfo_screenwidth()
screenheight = root.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
root.geometry(alignstr)
root.resizable(width=False, height=False)

# Create the Enter ID label
id_label = tk.Label(root)
id_label["anchor"] = "center"
ft = tkFont.Font(family='Helvetica', size=60, weight = 'bold')
id_label["font"] = ft
id_label["fg"] = "#FF6C40"
id_label["justify"] = "center"
id_label["text"] = "ENTER HACKERRANK ID'S!"
id_label['bg'] = '#404445'
id_label.place(x=15, y=2, width=1100, height=131)


def on_entry_click(event):
    if entry.get("1.0", 'end-1c') == '   Enter Comma Separated values of HACKERRANK_CONTEST_ID\'s':
        entry.delete('1.0', tk.END)

# Create the input field #FF6C40
entry = tk.Text(root)
entry["borderwidth"] = "5px"
entry['background'] = "black"
ft = tkFont.Font(family='Times', size=25, weight="bold")
entry["font"] = ft
entry.insert('1.0', '   Enter Comma Separated values of HACKERRANK_CONTEST_ID\'s')
entry.bind("<FocusIn>", on_entry_click)
entry["fg"] = "#FFE33E"
entry["relief"] = "groove"
entry.place(x=20, y=120, width=1101, height=431)
entry["insertbackground"] = "#FFE33E"


# Create the Generate button
def button_enter(event):
    generate_btn.config(background='black')


def button_leave(event):
    generate_btn.config(background='maroon')


generate_btn = tk.Button(root)
generate_btn.bind('<Enter>', button_enter)
generate_btn.bind('<Leave>', button_leave)
generate_btn["background"] = "maroon"
ft = tkFont.Font(family='Times', size=40, weight='bold')
generate_btn["borderwidth"] = "7px"
generate_btn["font"] = ft
generate_btn["fg"] = "#FFE33E"
generate_btn["justify"] = "center"
generate_btn["relief"] = "groove"
generate_btn["text"] = "Generate Excel Sheets!"
generate_btn["command"] = GButton_486_command
generate_btn.place(x=60, y=570, width=1010, height=99)
root.iconbitmap('venv/logo.ico')
root.wm_protocol(name='WM_DELETE_WINDOW')

# Start the tkinter event loop
root.mainloop()
